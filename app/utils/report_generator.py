"""
Report generation utilities
"""
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
import io


def generate_pdf_report(report, user):
    """Generate PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph(f"Sales Report - {report.period_type.title()}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Farmer info
    info = Paragraph(f"Farmer: {user.name}<br/>Date: {report.period_date}", styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 12))
    
    # Data table
    data = [
        ['Metric', 'Value'],
        ['Total Orders', str(report.total_orders)],
        ['Units Sold', f"{report.total_units_sold:.2f}"],
        ['Gross Revenue', f"₹{report.gross_revenue:,.2f}"],
        ['Net Profit', f"₹{report.net_profit:,.2f}"],
        ['Profit Margin', f"{report.profit_margin:.2f}%"]
    ]
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer


def generate_excel_report(report, user):
    """Generate Excel report"""
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Headers
    ws['A1'] = 'Sales Report'
    ws['A2'] = f'Farmer: {user.name}'
    ws['A3'] = f'Period: {report.period_type.title()}'
    ws['A4'] = f'Date: {report.period_date}'
    
    # Data
    ws['A6'] = 'Metric'
    ws['B6'] = 'Value'
    
    ws['A7'] = 'Total Orders'
    ws['B7'] = report.total_orders
    
    ws['A8'] = 'Units Sold'
    ws['B8'] = report.total_units_sold
    
    ws['A9'] = 'Gross Revenue'
    ws['B9'] = report.gross_revenue
    
    ws['A10'] = 'Net Profit'
    ws['B10'] = report.net_profit
    
    ws['A11'] = 'Profit Margin (%)'
    ws['B11'] = report.profit_margin
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer
